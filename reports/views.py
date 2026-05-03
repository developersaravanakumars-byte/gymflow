from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from members.models import Member
from payments.models import Payment
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def get_branches(user):
    return user.get_accessible_branches()


@login_required
def report_home(request):
    branches = get_branches(request.user)
    today = timezone.now().date()
    expiry_threshold = today + timezone.timedelta(days=30)

    context = {
        'total_active': Member.objects.filter(branch__in=branches, is_archived=False).count(),
        'expiring_30': Member.objects.filter(
            branch__in=branches, is_archived=False,
            membership_expiry__gte=today, membership_expiry__lte=expiry_threshold
        ).count(),
        'total_revenue': Payment.objects.filter(
            member__branch__in=branches, status='paid'
        ).aggregate(total=Sum('final_amount'))['total'] or 0,
        'pending_count': Payment.objects.filter(
            member__branch__in=branches, status__in=['pending', 'overdue']
        ).count(),
    }
    return render(request, 'reports/home.html', context)


@login_required
def revenue_report(request):
    branches = get_branches(request.user)
    month = request.GET.get('month', timezone.now().strftime('%Y-%m'))

    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        year, mon = timezone.now().year, timezone.now().month

    payments = Payment.objects.filter(
        member__branch__in=branches,
        created_at__year=year,
        created_at__month=mon,
    ).select_related('member', 'plan')

    total = payments.filter(status='paid').aggregate(t=Sum('final_amount'))['t'] or 0
    by_branch = payments.filter(status='paid').values(
        'member__branch__name'
    ).annotate(total=Sum('final_amount')).order_by('-total')

    return render(request, 'reports/revenue.html', {
        'payments': payments,
        'total': total,
        'by_branch': by_branch,
        'month': month,
    })


@login_required
def expiry_report(request):
    branches = get_branches(request.user)
    today = timezone.now().date()
    days = int(request.GET.get('days', 30))
    threshold = today + timezone.timedelta(days=days)

    members = Member.objects.filter(
        branch__in=branches,
        is_archived=False,
        membership_expiry__gte=today,
        membership_expiry__lte=threshold,
    ).order_by('membership_expiry').select_related('branch', 'plan')

    return render(request, 'reports/expiry.html', {
        'members': members,
        'days': days,
        'today': today,
    })


@login_required
def pending_report(request):
    branches = get_branches(request.user)
    payments = Payment.objects.filter(
        member__branch__in=branches,
        status__in=['pending', 'overdue'],
    ).select_related('member', 'member__branch').order_by('coverage_to')

    total_pending = payments.aggregate(t=Sum('final_amount'))['t'] or 0

    return render(request, 'reports/pending.html', {
        'payments': payments,
        'total_pending': total_pending,
    })


@login_required
def export_members_excel(request):
    branches = get_branches(request.user)
    members = Member.objects.filter(
        branch__in=branches, is_archived=False
    ).select_related('branch', 'plan').order_by('branch__name', 'first_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='7F77DD')
    header_align = Alignment(horizontal='center')

    headers = [
        'Name', 'Phone', 'Email', 'Branch', 'Plan',
        'Membership Start', 'Membership Expiry', 'Status', 'Joined'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    today = timezone.now().date()
    for row, member in enumerate(members, 2):
        status = 'Active' if member.membership_expiry and member.membership_expiry >= today else 'Expired'
        ws.append([
            member.full_name,
            member.phone,
            member.email,
            member.branch.name,
            member.plan.name if member.plan else '',
            member.membership_start,
            member.membership_expiry,
            status,
            member.joined_at.date(),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="gymflow_members.xlsx"'
    wb.save(response)
    return response
